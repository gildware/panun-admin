#!/usr/bin/env python3
"""Export full active catalog Excel: one sheet per main category."""

from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/Users/kamran/Desktop/panun kaergar/panun-admin/scripts/backups") / (
    "Panun-Kaergar-Active-Catalog-" + __import__("datetime").datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
)
PHP_EXPORT = Path("/tmp/pk-active-catalog-export.json")


def fetch_rows() -> list[dict]:
    php = r'''
require "vendor/autoload.php";
$app = require "bootstrap/app.php";
$app->make(Illuminate\Contracts\Console\Kernel::class)->bootstrap();
config(["database.connections.live" => [
  "driver" => "mysql",
  "host" => "82.25.121.201",
  "port" => "3306",
  "database" => "u397782854_live_pk_dec",
  "username" => "u397782854_live_pk_usr",
  "password" => getenv("LIVE_DB_PASSWORD"),
  "charset" => "utf8mb4",
  "collation" => "utf8mb4_unicode_ci",
  "prefix" => "",
  "strict" => true,
]]);
$db = Illuminate\Support\Facades\DB::connection("live");
$rows = $db->select("
  SELECT
    cat.name AS category_name,
    cat.slug AS category_slug,
    COALESCE(cat.sort_order, 999) AS category_sort,
    COALESCE(sub.name, '') AS subcategory_name,
    COALESCE(sub.slug, '') AS subcategory_slug,
    COALESCE(sub.sort_order, 0) AS subcategory_sort,
    s.name AS service_name,
    s.slug AS service_slug,
    s.sort_order AS service_sort,
    sv.title AS variation_name,
    COALESCE(NULLIF(TRIM(sv.description), ''), '') AS variation_description,
    sv.variant_key,
    sv.sort_order AS variation_sort,
    sv.note AS variation_note
  FROM service_variants sv
  INNER JOIN services s ON s.id = sv.service_id
  INNER JOIN categories cat ON cat.id = s.category_id
  LEFT JOIN categories sub ON sub.id = s.sub_category_id
  WHERE s.is_active = 1
    AND s.deleted_at IS NULL
    AND sv.is_active = 1
  ORDER BY category_sort, category_name, subcategory_sort, subcategory_name, service_sort, s.name, variation_sort, sv.title
");
$out = [];
foreach ($rows as $r) {
  $out[] = [
    "category_name" => (string) $r->category_name,
    "category_slug" => (string) $r->category_slug,
    "subcategory_name" => (string) $r->subcategory_name,
    "service_name" => (string) $r->service_name,
    "variation_name" => (string) $r->variation_name,
    "variation_description" => (string) $r->variation_description,
    "variation_note" => (string) ($r->variation_note ?? ""),
  ];
}
file_put_contents("/tmp/pk-active-catalog-export.json", json_encode($out, JSON_UNESCAPED_UNICODE));
echo count($out);
'''
    env = os.environ.copy()
    if "LIVE_DB_PASSWORD" not in env or not env["LIVE_DB_PASSWORD"]:
        raise SystemExit("Set LIVE_DB_PASSWORD")
    result = subprocess.run(
        ["php", "-r", php],
        cwd="/Users/kamran/Desktop/panun kaergar/panun-admin",
        env=env,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise SystemExit(result.stderr or result.stdout)
    print("Fetched", result.stdout.strip(), "rows")
    return json.loads(PHP_EXPORT.read_text())


def sheet_title(name: str, used: set[str]) -> str:
    # Excel sheet name max 31 chars, no: \ / ? * [ ]
    clean = re.sub(r'[\\/*?:\[\]]', "", name).strip() or "Category"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in used:
        suffix = f" ({i})"
        clean = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(clean)
    return clean


def build_workbook(rows: list[dict]) -> Path:
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    by_cat: dict[str, list[dict]] = {}
    cat_order: list[str] = []
    for r in rows:
        key = r["category_name"] or r["category_slug"] or "Uncategorized"
        if key not in by_cat:
            by_cat[key] = []
            cat_order.append(key)
        by_cat[key].append(r)

    headers = [
        "Sub Category",
        "Service Name",
        "Variation",
        "Variation Description",
        "Handled by",
        "Provider one pricing",
        "Provider two pricing",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A233A")
    thin = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    used_titles: set[str] = set()
    for cat_name in cat_order:
        items = by_cat[cat_name]
        ws = wb.create_sheet(title=sheet_title(cat_name, used_titles))
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = thin

        for i, r in enumerate(items, 2):
            desc = r["variation_description"] or r.get("variation_note") or ""
            values = [
                r["subcategory_name"] or "",
                r["service_name"],
                r["variation_name"],
                desc,
                "",  # Handled by
                "",  # Provider one pricing
                "",  # Provider two pricing
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(i, col, val)
                cell.alignment = wrap
                cell.border = thin

        widths = [28, 36, 42, 55, 18, 22, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{len(items) + 1}"

    # Summary sheet first
    summary = wb.create_sheet(title="Summary", index=0)
    summary["A1"] = "Category"
    summary["B1"] = "Active variations"
    summary["C1"] = "Active services"
    for cell in (summary["A1"], summary["B1"], summary["C1"]):
        cell.font = header_font
        cell.fill = header_fill
    for i, cat_name in enumerate(cat_order, 2):
        items = by_cat[cat_name]
        services = {x["service_name"] for x in items}
        summary.cell(i, 1, cat_name)
        summary.cell(i, 2, len(items))
        summary.cell(i, 3, len(services))
    summary.cell(len(cat_order) + 2, 1, "TOTAL")
    summary.cell(len(cat_order) + 2, 2, len(rows))
    summary.cell(len(cat_order) + 2, 3, len({(r["category_name"], r["service_name"]) for r in rows}))
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 16

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


def main() -> None:
    rows = fetch_rows()
    path = build_workbook(rows)
    cats = sorted({r["category_name"] for r in rows})
    print(f"Wrote {path}")
    print(f"Categories: {len(cats)}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
